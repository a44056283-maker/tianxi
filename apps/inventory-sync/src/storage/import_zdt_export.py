#!/usr/bin/env python3
"""
将ZDT导出的Excel文件数据导入到zdt_sync PostgreSQL数据库

导出文件：
  orderData.xlsx       → 订单主表（门店/收银员/订单状态）
  orderProductData.xlsx → 订单明细（商品名称/MTM/规格/SN）

用法：
  python3 import_zdt_export.py <orderData.xlsx> <orderProductData.xlsx>
"""
import sys, os, json, subprocess, psycopg2
from datetime import datetime

DB = dict(host='localhost', port=5432, dbname='zdt_sync', user='zdt', password='zdt')

def read_xlsx(path):
    code = f'''
import openpyxl, json
wb = openpyxl.load_workbook("{path}", data_only=True)
ws = wb.active
headers = [str(c.value or '') for c in ws[1]]
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if any(v is not None for v in row):
        rows.append(dict(zip(headers, [str(v or '') for v in row])))
print(json.dumps({{'headers': headers, 'rows': rows}}, ensure_ascii=False))
'''
    out = subprocess.check_output(['python3', '-c', code], encoding='utf-8')
    return json.loads(out)

def get_conn():
    return psycopg2.connect(**DB)

def upsert_fact_orders(conn, rows):
    """更新 fact_orders.shop_name, status_name, cashier_name"""
    cur = conn.cursor()
    updated = 0
    notfound = 0
    for row in rows:
        order_no = row.get('订单号','').strip()
        if not order_no:
            continue
        shop_name = row.get('门店','').strip()
        status_name = row.get('订单状态','').strip()
        cashier_name = row.get('收银员','').strip()
        
        cur.execute("""
            UPDATE fact_orders 
            SET shop_name = %s, status_name = %s, cashier_name = %s
            WHERE order_no = %s
            RETURNING order_no
        """, (shop_name, status_name, cashier_name, order_no))
        if cur.fetchone():
            updated += 1
        else:
            notfound += 1
    conn.commit()
    return updated, notfound

def upsert_fact_order_items(conn, rows):
    """更新 fact_order_items.product_name, mtm_code, spec"""
    cur = conn.cursor()
    updated = 0
    notfound = 0
    for row in rows:
        order_no = row.get('订单号','').strip()
        if not order_no:
            continue
        
        # 先找 order_id
        cur.execute("SELECT order_id FROM fact_orders WHERE order_no = %s", (order_no,))
        res = cur.fetchone()
        if not res:
            notfound += 1
            continue
        order_id = res[0]

        product_name = row.get('商品信息','').strip()
        mtm_code = row.get('PN/MTM','').strip()
        spec = row.get('商品规格','').strip()
        sn = row.get('商品SN','').strip()

        # 更新 fact_order_items
        cur.execute("""
            UPDATE fact_order_items 
            SET product_name = %s, mtm_code = %s, spec = %s
            WHERE order_id = %s
            RETURNING id
        """, (product_name, mtm_code, spec, order_id))
        if cur.fetchone():
            updated += 1
        else:
            notfound += 1
        
        # 更新 SN (通过 serial_number)
        if sn:
            cur.execute("""
                UPDATE fact_order_items
                SET serial_number = %s
                WHERE order_id = %s AND (serial_number = '' OR serial_number IS NULL)
                RETURNING id
            """, (sn, order_id))
    conn.commit()
    return updated, notfound

def update_sync_state(conn):
    """更新 sync_state 时间戳"""
    cur = conn.cursor()
    now = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
    cur.execute("""
        UPDATE sync_state 
        SET last_sync_time = %s, last_success_time = %s, status = 'active'
        WHERE entity_name IN ('orders_offline', 'orders_online')
    """, (now, now))
    conn.commit()
    return now

def main():
    if len(sys.argv) < 3:
        print('用法: python3 import_zdt_export.py <orderData.xlsx> <orderProductData.xlsx>')
        sys.exit(1)

    order_file = sys.argv[1]
    product_file = sys.argv[2]

    if not os.path.exists(order_file):
        print(f"文件不存在: {order_file}")
        sys.exit(1)
    if not os.path.exists(product_file):
        print(f"文件不存在: {product_file}")
        sys.exit(1)

    print(f"读取订单主表: {order_file}")
    order_data = read_xlsx(order_file)
    print(f"  字段: {order_data['headers']}")
    print(f"  数据: {len(order_data['rows'])} 条")

    print(f"\n读取订单明细: {product_file}")
    product_data = read_xlsx(product_file)
    print(f"  字段: {product_data['headers']}")
    print(f"  数据: {len(product_data['rows'])} 条")

    conn = get_conn()
    try:
        print("\n=== 更新 fact_orders ===")
        u, n = upsert_fact_orders(conn, order_data['rows'])
        print(f"  更新: {u} 条, 未找到: {n} 条")

        print("\n=== 更新 fact_order_items ===")
        u2, n2 = upsert_fact_order_items(conn, product_data['rows'])
        print(f"  更新: {u2} 条, 未找到: {n2} 条")

        print("\n=== 更新 sync_state ===")
        ts = update_sync_state(conn)
        print(f"  完成: {ts}")

        # 验证结果
        cur = conn.cursor()
        cur.execute("""
            SELECT 
                COUNT(*) as total,
                COUNT(status_name) FILTER (WHERE status_name != '') as status_ok,
                COUNT(shop_name) FILTER (WHERE shop_name != '') as shop_ok
            FROM fact_orders WHERE order_no LIKE 'XS%%'
        """)
        r = cur.fetchone()
        print(f"\n=== 验证 fact_orders (XS订单) ===")
        print(f"  总数: {r[0]}, status_name非空: {r[1]}, shop_name非空: {r[2]}")

    finally:
        conn.close()

    print("\n导入完成!")

if __name__ == '__main__':
    main()
